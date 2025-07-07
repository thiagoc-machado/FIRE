from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiExample, OpenApiResponse
from rest_framework.parsers import MultiPartParser
import os
import subprocess
from django.http import FileResponse
from django.conf import settings
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from django.utils.timezone import now
from decimal import Decimal
from .models import Category, Broker, Asset, Operation
from .serializers import CategorySerializer, BrokerSerializer, AssetSerializer, OperationSerializer
from .services.yahoo_service import buscar_dados_ativo
from django.http import JsonResponse
from django.db.models.functions import TruncMonth
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from datetime import date
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from core.serializers import FireCalculatorInputSerializer
from calendar import monthrange

@extend_schema(tags=['Categorias'])
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


@extend_schema(tags=['Corretoras'])
class BrokerViewSet(viewsets.ModelViewSet):
    queryset = Broker.objects.all()
    serializer_class = BrokerSerializer
    permission_classes = [IsAuthenticated]


@extend_schema(tags=['Ativos'])
class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary='Atualizar cotações via Yahoo Finance',
        description='Busca e retorna a cotação atual, moeda, dividend yield e frequência de dividendos para todos os ativos cadastrados, usando a API do Yahoo Finance.',
        responses={
            200: OpenApiResponse(
                description='Lista de ativos com dados atualizados',
                examples=[
                    OpenApiExample(
                        'Exemplo de resposta',
                        value={
                            'ativos_atualizados': [
                                {
                                    'codigo': 'AAPL',
                                    'nome': 'Apple Inc.',
                                    'cotacao': 213.55,
                                    'moeda': 'USD',
                                    'dividend_yield': 0.51,
                                    'frequencia': 1.04
                                }
                            ]
                        },
                        response_only=True
                    )
                ]
            )
        }
    )
    @action(detail=False, methods=['get'], url_path='atualizar')
    def atualizar_cotacoes(self, request):
        atualizados = []
        for asset in self.get_queryset():
            dados = buscar_dados_ativo(asset.codigo)

            try:
                cotacao = float(dados['cotacao'])
            except (ValueError, TypeError):
                cotacao = None

            if cotacao:
                asset.preco_atual = cotacao
                asset.moeda = dados.get('moeda', asset.moeda)
                asset.frequencia_dividendos = dados.get('frequencia', asset.frequencia_dividendos)
                asset.save()

                atualizados.append({
                    'codigo': asset.codigo,
                    'nome': dados.get('nome', asset.nome),
                    'cotacao': asset.preco_atual,
                    'moeda': asset.moeda,
                    'dividend_yield': dados.get('dividend_yield'),
                    'frequencia': asset.frequencia_dividendos,
                })

        return Response({'ativos_atualizados': atualizados})
    
    @action(detail=False, methods=['get'], url_path='buscar')
    def buscar_ativo(self, request):
        codigo = request.query_params.get('codigo')
        if not codigo:
            return Response({'detail': 'Código não informado.'}, status=400)

        try:
            dados = buscar_dados_ativo(codigo.upper())
            print(f'Dados brutos recebidos do Yahoo para {codigo.upper()}: {dados}') 
        except Exception as e:
            return Response({'detail': f'Erro ao buscar dados: {str(e)}'}, status=500)

        if not dados:
            return Response({'detail': 'Dados não encontrados para este código.'}, status=404)

        try:
            cotacao = float(dados.get('cotacao'))
        except (ValueError, TypeError):
            return Response({'detail': 'Cotação inválida.'}, status=404)

        return Response({
            'codigo': codigo.upper(),
            'nome': dados.get('nome'),
            'moeda': dados.get('moeda'),
            'preco_atual': round(cotacao, 2),
            'frequencia_dividendos': dados.get('frequencia'),
            'dividend_yield': dados.get('dividend_yield'),
            'categoria': dados.get('categoria'),
            'tipo': dados.get('tipo'),
            'moeda': dados.get('moeda')
        })

@extend_schema(tags=['Operações'])
class OperationViewSet(viewsets.ModelViewSet):
    queryset = Operation.objects.none()
    serializer_class = OperationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Operation.objects.filter(user=self.request.user)

    @extend_schema(
        summary='Criar operação de compra ou venda',
        examples=[
            OpenApiExample(
                'Exemplo de operação',
                value={
                    'tipo': 'COMPRA',
                    'ativo': 1,
                    'quantidade': 10,
                    'valor_compra': 150.0,
                    'data': '2024-01-01',
                    'dividendos': 5.0,
                    'categoria': 1,
                    'corretora': 1
                },
                request_only=True
            )
        ]
    )
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @extend_schema(
        summary='Importar operações via CSV',
        description='Permite importar operações de compra/venda a partir de um arquivo CSV com cabeçalhos: tipo,ativo,quantidade,valor_compra,data,dividendos,categoria,corretora.',
        methods=['POST'],
        request=None,
        responses={201: OpenApiResponse(description='Importação realizada com sucesso')}
    )
    @action(detail=False, methods=['post'], url_path='importar', parser_classes=[MultiPartParser])
    def importar(self, request):
        import csv, io
        file = request.FILES.get('file')
        if not file:
            return Response({'erro': 'Arquivo não enviado'}, status=status.HTTP_400_BAD_REQUEST)

        decoded_file = file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)

        criadas = 0
        for row in reader:
            try:
                data = {
                    'user': request.user.id,
                    'tipo': row['tipo'],
                    'ativo': int(row['ativo']),
                    'quantidade': float(row['quantidade']),
                    'valor_compra': float(row['valor_compra']),
                    'data': row['data'],
                    'dividendos': float(row.get('dividendos', 0)),
                    'categoria': int(row['categoria']),
                    'corretora': int(row['corretora']),
                }
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                serializer.save(user=request.user)
                criadas += 1
            except Exception:
                continue

        return Response({'mensagem': f'{criadas} operações importadas com sucesso.'}, status=201)

    @extend_schema(
        summary='Exportar operações do usuário',
        description='Retorna todas as operações do usuário autenticado em formato JSON.',
        responses={200: OpenApiResponse(description='Lista de operações exportadas')}
    )
    @action(detail=False, methods=['get'], url_path='exportar')
    def exportar(self, request):
        operacoes = self.get_queryset()
        serializer = self.get_serializer(operacoes, many=True)
        return JsonResponse(serializer.data, safe=False)
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from io import BytesIO

        user = request.user
        operacoes = self.get_queryset().select_related('ativo', 'categoria', 'corretora')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Operações'

        cabecalhos = [
            'Usuário', 'Tipo', 'Ativo', 'Quantidade', 'Valor de Compra', 'Data',
            'Dividendos', 'Categoria', 'Corretora'
        ]
        ws.append(cabecalhos)

        for op in operacoes:
            ws.append([
                user.email,
                op.tipo,
                op.ativo.nome if op.ativo else '',
                float(op.quantidade),
                float(op.valor_compra),
                op.data.strftime('%Y-%m-%d'),
                float(op.dividendos),
                op.categoria.nome if op.categoria else '',
                op.corretora.nome if op.corretora else ''
            ])

        # Largura automática
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'operacoes_{user.username}.xlsx'
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
class BackupViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminUser]

    @action(detail=False, methods=['get'], url_path='backup')
    def backup(self, request):
        user = request.user
        filename = f'fire_backup_{user.username}.sql.gz'
        tmp = tempfile.NamedTemporaryFile(suffix='.sql.gz', delete=False)

        db = settings.DATABASES['default']
        cmd = [
            'pg_dump',
            '-U', db['USER'],
            '-h', db.get('HOST', 'localhost'),
            '-d', db['NAME'],
            '--no-owner',
            '--no-acl'
        ]

        env = os.environ.copy()
        env['PGPASSWORD'] = db['PASSWORD']

        gzip = subprocess.Popen(['gzip'], stdin=subprocess.PIPE, stdout=tmp)
        dump = subprocess.Popen(cmd, stdout=gzip.stdin, env=env)
        dump.wait()
        gzip.stdin.close()
        gzip.wait()
        tmp.seek(0)

        return FileResponse(tmp, as_attachment=True, filename=filename)

@extend_schema(tags=['Dashboard'])
class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary='Totais do dashboard',
        description='Retorna o total investido, valorização e dividendos recebidos do usuário.',
        responses={
            200: OpenApiResponse(
                description='Totais agregados',
                examples=[
                    OpenApiExample(
                        'Exemplo',
                        value={'investido': 15000.0, 'valorizacao': 1200.0, 'dividendos': 350.0}
                    )
                ]
            )
        }
    )
    @action(detail=False, methods=['get'], url_path='total')
    def total(self, request):
        user = request.user
        qs = Operation.objects.filter(user=user)

        total_investido = qs.aggregate(
            total=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
        )['total'] or 0

        total_dividendos = qs.aggregate(total=Sum('dividendos'))['total'] or 0

        # Valorização: soma de (preco_atual - valor_compra) * quantidade
        qs_valorizacao = qs.select_related('ativo')
        total_valorizacao = 0
        for op in qs_valorizacao:
            print(f'op.ativo: {op.ativo}, op.ativo.preco_atual: {op.ativo.preco_atual}')
            if op.ativo and op.ativo.preco_atual:
                total_valorizacao += (op.ativo.preco_atual - op.valor_compra) * op.quantidade

        return Response({
            'investido': round(total_investido, 2),
            'valorizacao': round(total_valorizacao, 2),
            'dividendos': round(total_dividendos, 2)
        })

    @action(detail=False, methods=['get'], url_path='por-categoria')
    def por_categoria(self, request):
        user = request.user
        categorias = Category.objects.filter(operation__user=user).distinct()
        resultado = []
        for cat in categorias:
            total = Operation.objects.filter(user=user, categoria=cat).aggregate(
                valor=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
            )['valor'] or 0
            resultado.append({'nome': cat.nome, 'valor': round(total, 2)})
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='por-corretora')
    def por_corretora(self, request):
        user = request.user
        corretoras = Broker.objects.filter(operation__user=user).distinct()
        resultado = []
        for broker in corretoras:
            total = Operation.objects.filter(user=user, corretora=broker).aggregate(
                valor=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
            )['valor'] or 0
            resultado.append({'nome': broker.nome, 'valor': round(total, 2)})
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='por-ativo')
    def por_ativo(self, request):
        user = request.user
        ativos = Asset.objects.filter(operation__user=user).distinct()
        resultado = []
        for ativo in ativos:
            total = Operation.objects.filter(user=user, ativo=ativo).aggregate(
                valor=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
            )['valor'] or 0
            if total > 0:
                resultado.append({'codigo': ativo.codigo, 'valor': round(total, 2)})
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='dividendos-mensais')
    def dividendos_mensais(self, request):
        user = request.user
        qs = Operation.objects.filter(user=user).annotate(mes=TruncMonth('data')).values('mes').annotate(
            total=Sum('dividendos')
        ).order_by('mes')
        resultado = [{'mes': r['mes'].strftime('%Y-%m'), 'total': round(r['total'] or 0, 2)} for r in qs]
        return Response(resultado)
    
    @action(detail=False, methods=['get'], url_path='progresso-fire')
    def progresso_fire(self, request):
        user = request.user
        operations = Operation.objects.filter(user=user).select_related('ativo')

        # Total investido
        investido = operations.aggregate(
            total=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
        )['total'] or 0

        # Valorização
        valorizacao = 0
        for op in operations:
            if op.ativo and op.ativo.preco_atual:
                valorizacao += (op.ativo.preco_atual - op.valor_compra) * op.quantidade


        patrimonio = Decimal(investido) + valorizacao

        # Dividendos médios mensais (últimos 12 meses)
        from datetime import timedelta

        inicio = now() - timedelta(days=365)
        dividendos_ano = Operation.objects.filter(user=user, data__gte=inicio).annotate(
            mes=TruncMonth('data')
        ).values('mes').annotate(total=Sum('dividendos')).order_by()

        total_dividendos = sum(d['total'] or 0 for d in dividendos_ano)
        media_dividendo_mensal = total_dividendos / (len(dividendos_ano) or 1)

        # Aportes médios mensais (baseado em valor_compra * quantidade)
        aportes_ano = Operation.objects.filter(user=user, tipo='COMPRA', data__gte=inicio).annotate(
            mes=TruncMonth('data')
        ).values('mes').annotate(
            total=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
        )

        total_aportes = sum(a['total'] or 0 for a in aportes_ano)
        media_aporte_mensal = total_aportes / (len(aportes_ano) or 1)

        # Configurações do usuário (simulado por enquanto)
        settings = getattr(user, 'settings', None)

        meta_total = float(settings.meta_fire_total) if settings else 600000
        renda_desejada = float(settings.renda_fire_desejada) if settings else 2000

        percentual = (patrimonio / Decimal(meta_total)) * 100 if meta_total else 0
        gap_mensal = renda_desejada - media_dividendo_mensal
        meses_estimados = (Decimal(meta_total) - patrimonio) / (Decimal(media_aporte_mensal) or 1)

        # patrimonio = Decimal(investido) + valorizacao

        return Response({
            'meta_total': round(meta_total, 2),
            'patrimonio_atual': round(patrimonio, 2),
            'percentual_atingido': round(percentual, 2),
            'renda_desejada': round(renda_desejada, 2),
            'renda_atual_estimada': round(media_dividendo_mensal, 2),
            'gap_mensal': round(gap_mensal, 2),
            'meses_estimados': int(meses_estimados),
            'aporte_mensal_medio': round(media_aporte_mensal, 2)
        })

    @action(detail=False, methods=['get'], url_path='evolucao')
    def evolucao(self, request):
        user = request.user
        qs = Operation.objects.filter(user=user).annotate(mes=TruncMonth('data'))

        resultado = []

        dados_por_mes = qs.values('mes').annotate(
            investido=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
        ).order_by('mes')

        for item in dados_por_mes:
            mes = item['mes'].strftime('%Y-%m')
            investido = item['investido'] or 0

            # Valorização real: (preco_atual - valor_compra) * quantidade
            valorizado = 0
            for op in qs.filter(data__month=item['mes'].month, data__year=item['mes'].year).select_related('ativo'):
                if op.ativo and op.ativo.preco_atual:
                    valorizado += float(op.quantidade) * float(op.ativo.preco_atual)

            resultado.append({
                'mes': mes,
                'investido': round(investido, 2),
                'valorizado': round(valorizado, 2),
            })

        return Response(resultado)
    
    @action(detail=False, methods=['get'], url_path='planejamento')
    def planejamento(self, request):
        user = request.user
        qs = Operation.objects.filter(user=user).annotate(mes=TruncMonth('data'))

        # Aportes reais por mês
        dados_reais = qs.values('mes').annotate(
            aporte_real=Sum(F('valor_compra') * F('quantidade'), output_field=FloatField())
        ).order_by('mes')

        # Meta mensal vinda do UserSettings
        try:
            settings = user.settings
            aporte_planejado = float(settings.meta_fire_total or 0) / 300  # ex: 25 anos = 300 meses
        except Exception:
            aporte_planejado = 0

        resultado = [{
            'mes': r['mes'].strftime('%Y-%m'),
            'aporte_real': round(r['aporte_real'] or 0, 2),
            'aporte_planejado': round(aporte_planejado, 2)
        } for r in dados_reais]

        return Response(resultado)

class FireCalculatorView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = FireCalculatorInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        aporte = data['aporte_mensal']
        patrimonio = data['patrimonio_atual']
        meta = data['meta_fire_total']
        juros_mensal = (1 + data['rentabilidade_anual'] / 100) ** (1/12) - 1

        meses = 0
        labels = []
        valores = []
        hoje = date.today()

        while patrimonio < meta and meses < 1000:
            labels.append(f'{hoje.year}-{hoje.month:02d}')
            valores.append(round(patrimonio, 2))

            patrimonio = patrimonio * (1 + juros_mensal) + aporte
            meses += 1

            # avança para o próximo mês
            if hoje.month == 12:
                hoje = date(hoje.year + 1, 1, 1)
            else:
                hoje = date(hoje.year, hoje.month + 1, 1)

        return Response({
            'meses_estimados': meses,
            'data_prevista': hoje.strftime('%Y-%m-%d'),
            'grafico': {
                'labels': labels,
                'valores': valores,
            }
        })